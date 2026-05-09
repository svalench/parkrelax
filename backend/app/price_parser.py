import io
from typing import Any

import openpyxl


def _is_price(text: Any) -> bool:
    if not text:
        return False
    s = str(text).strip().lower()
    return "руб" in s and any(ch.isdigit() for ch in s)


def parse_price_excel(file_content: bytes) -> list[dict[str, Any]]:
    """Parse price-list Excel.

    Heuristic:
      - col_a present, col_b empty or not a price  -> category header
      - col_a present, col_b is a price             -> price item
      - col_a starts with *                         -> note
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_content))
    ws = wb.active

    current_category: str | None = None
    items: list[dict[str, Any]] = []

    for row in ws.iter_rows(values_only=True):
        col_a = str(row[0]).strip() if row[0] else None
        col_b = str(row[1]).strip() if len(row) > 1 and row[1] else None

        if not col_a and not col_b:
            continue

        # Note row
        if col_a and col_a.startswith("*"):
            items.append(
                {
                    "category": current_category,
                    "name": col_a,
                    "price": "",
                    "isNote": True,
                }
            )
            continue

        # Category detection: col_a has text, col_b is empty or not a price
        if col_a and (not col_b or not _is_price(col_b)):
            current_category = col_a
            continue

        # Price item
        if col_a and col_b and _is_price(col_b) and current_category:
            items.append(
                {
                    "category": current_category,
                    "name": col_a,
                    "price": col_b,
                    "isNote": False,
                }
            )

    return items


def generate_template_excel() -> bytes:
    """Generate a blank price-list template."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Прайс-лист"

    ws["A1"] = "Категория / Наименование"
    ws["B1"] = "Цена"
    ws["A2"] = "ПРОЖИВАНИЕ"
    ws["A3"] = "Коттедж до 8 человек"
    ws["B3"] = "550 руб"
    ws["A4"] = "Апартаменты №1 до 3 человек"
    ws["B4"] = "230 руб"
    ws["A5"] = "АРЕНДА"
    ws["A6"] = "Беседка на пляже (на 8 мест)"
    ws["B6"] = "50 руб за 4 часа"
    ws["A7"] = "*Примечание: цены указаны на 2026 год"

    # Adjust column widths
    ws.column_dimensions["A"].width = 60
    ws.column_dimensions["B"].width = 30

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
